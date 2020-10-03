from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt

#________________________________________________________________________________________________________
#load data from excel file
def load_xlsx(path, x_coord, y_coord):
    wb = load_workbook(path)
    sheet = wb.active
    if type(x_coord) == str and type(y_coord) == int:
        values = np.array([i.value for i in sheet[x_coord]][y_coord::])
        return values

    return print("InputError_LoadFunction")


#________________________________________________________________________________________________________
#load data from two columns in an excel file and combine them
def load_multicolumn_xlsx(path, x_coord_1, x_coord_2, y_coord):
    wb = load_workbook(path)
    sheet = wb.active
    if type(x_coord_1)==str and type(x_coord_2)==str and type(y_coord) == int:
        coord_tuple = (x_coord_1,x_coord_2)       
        values_col_1 = np.array([i.value for i in sheet[x_coord_1]][y_coord::])
        values_col_2 = np.array([i.value for i in sheet[x_coord_2]][y_coord::])
        values = np.array([str(values_col_1[j])+str(values_col_2[j]) for j in range(len(values_col_1))])       
        return values
    return print("InputError_LoadFunction")
    

#________________________________________________________________________________________________________
#replace the unavailable datapoints with the last available datapoint
def NA_convert(array,notavailable = "NA"):
    for _ in range(len(array)):
        if array[_] == notavailable:
            array[_] = array[_-1]
    return array


#________________________________________________________________________________________________________
#convert the data from string format to the float format
def string_convert(array):
    return array.astype(np.float)
 
#________________________________________________________________________________________________________
#moving average process

def mov_avg(x, N):
    cumsum = np.cumsum(np.insert(x, 0, 0)) 
    raw = (cumsum[N:] - cumsum[:-N]) / float(N)
    for i in range(math.floor(int(N/2))):
        raw = np.insert(raw,i,x[i])
    for i in range(math.floor(int(N/2)-1),0,-1):
        raw = np.append(raw,x[-i])
    return raw


